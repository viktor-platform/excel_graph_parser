import os
from io import BytesIO
from pathlib import Path
from typing import List, Dict, Union

from munch import Munch
from openpyxl import load_workbook
from openpyxl.cell import Cell
from viktor import UserError, UserMessage
from viktor.errors import InputViolation
from viktor.external.spreadsheet import SpreadsheetCalculationInput, SpreadsheetCalculation

import plotly.graph_objects as go

ALLOWED_FIGURE_TYPES = ["lineChart", "scatterChart", "barChart", "pieChart"]


class ExcelImageParser:
    def __init__(self, excel_file_path: Union[Path, str], params: Munch, from_app: bool = False):
        self.workbook = load_workbook(filename=excel_file_path, data_only=True)
        self.params = params
        self.excel_file_path = excel_file_path
        self.from_app = from_app

        # Add exit point name to dataframe
        sheet_names = self.workbook.sheetnames

        # Loop through sheets
        self.sheets = []
        self.charts = []

        for sheet_name in sheet_names:
            self.sheets.append(self.workbook[sheet_name])

        for sheet in self.sheets:
            # Check if there are charts in the sheet
            if sheet._charts:
                # Loop through each chart in the sheet
                for chart in sheet._charts:
                    self.charts.append(chart)

    def get_input_cells(self) -> List[Dict]:
        """Gets inputs from the excel file as a dict"""
        wb = self.workbook
        ws_input = wb["viktor-input-sheet"]
        inputs = []
        for index, row in enumerate(ws_input.iter_rows(min_row=2, max_col=4)):
            if row[0].value:
                inputs.append(
                    {
                        "name": row[0].value,
                        "unit": row[1].value if row[1].value else "",
                        "description": row[2].value,
                        "default": row[3].value,
                        "key": f"input_{index}",
                    }
                )
        return inputs

    def get_evaluated_spreadsheet(self):
        """Evaluate spreadsheet so the version with the updated inputs and outputs is available"""
        inputs = []
        input_cells = self.get_input_cells()

        # Check whether the user wrongfully adjusted the inputs table
        if not self.from_app:
            if len(input_cells) != len(self.params.preview_step.fields_table):
                raise UserError(
                    "Please do not add or delete rows from the input table, go back to the previous step and re-process"
                    " the uploaded file"
                )

        # Load spreadsheet with correct inputs
        if not self.from_app:
            for (row, input_cell) in zip(self.params.preview_step.fields_table, input_cells):
                field_name = input_cell["name"]
                value = row["values"]
                inputs.append(SpreadsheetCalculationInput(field_name, value))

            spreadsheet = SpreadsheetCalculation(self.params.upload_step.excel_file.file, inputs)

        else:
            for input_cell in input_cells:
                field_name = input_cell["name"]
                value = self.params[input_cell["key"]]
                inputs.append(SpreadsheetCalculationInput(field_name, value))

            spreadsheet = SpreadsheetCalculation.from_path(self.excel_file_path, inputs)
        result = spreadsheet.evaluate(include_filled_file=True)
        evaluated_workbook = load_workbook(BytesIO(result.file_content), data_only=True)

        return evaluated_workbook, result

    def get_outputs(self) -> List[Dict]:
        """Gets outputs from the excel file as a dict (will return empty if no outputs are present in sheet)"""
        wb, result = self.get_evaluated_spreadsheet()
        values = result.values
        ws_output = wb["viktor-output-sheet"]
        outputs = []
        if not values:
            return outputs
        for index, row in enumerate(ws_output.iter_rows(min_row=2, max_col=4)):
            name = row[0].value
            unit = row[1].value if row[1].value else ""
            description = row[2].value
            if name:
                outputs.append(
                    {"name": name, "unit": unit, "description": description, "key": f"output_{index}",
                     "value": values[row[0].value], "type": str(type(values[row[0].value]))}
                )
        return outputs

    def get_figures_from_excel_file(self) -> list:
        """Gets figures from the excel file as a list"""

        wb, _ = self.get_evaluated_spreadsheet()
        figures = []

        for i, chart in enumerate(self.charts):
            # Get the general chart elements
            series = []
            if chart.title:
                chart_title = ''.join([title_element.t for title_element in chart.title.tx.rich.p[0].r])
            else:
                chart_title = f"Untitled Chart {i}"
            chart_type = chart.tagname
            if chart_type not in ALLOWED_FIGURE_TYPES:
                UserMessage.warning(f"Chart titled {chart_title} is not of one of the allowed types and can not be visualised")
                continue
            x_axis_title, y_axis_title = None, None
            if chart_type != "pieChart":
                if chart.x_axis.title:
                    x_axis_title = chart.x_axis.title.tx.rich.p[0].r[-1].t
                if chart.y_axis.title:
                    y_axis_title = chart.y_axis.title.tx.rich.p[0].r[-1].t

            # Get series data
            for i, serie in enumerate(chart.series):
                if chart_type == "scatterChart":
                    if serie.xVal:
                        if serie.xVal.strRef:
                            input_cat_range = serie.xVal.strRef.f
                            input_cat_format = None
                        elif serie.xVal.numRef:
                            input_cat_range = serie.xVal.numRef.f
                            input_cat_format = serie.xVal.numRef.numCache.formatCode
                            input_cat_format = None if input_cat_format == "General" else input_cat_format

                    input_val_range = serie.yVal.numRef.f
                    input_val_format = serie.yVal.numRef.numCache.formatCode
                    input_val_format = None if input_val_format == "General" else input_val_format

                else:
                    if serie.cat:
                        # if no category data in the sequence, use the one that was set for the previous sequence
                        if serie.cat.strRef:
                            input_cat_range = serie.cat.strRef.f
                            input_cat_format = None
                        elif serie.cat.numRef:
                            input_cat_range = serie.cat.numRef.f
                            input_cat_format = serie.cat.numRef.numCache.formatCode
                            input_cat_format = None if input_cat_format == "General" else input_cat_format

                    input_val_range = serie.val.numRef.f
                    input_val_format = serie.val.numRef.numCache.formatCode
                    input_val_format = None if input_val_format == "General" else input_val_format

                chart_sheet_name = input_cat_range.replace('(', '').replace(')', '').replace("'", "").split(sep="!")[0]

                chart_cat_range = input_cat_range.replace('(', '').replace(')', '').replace("'", "").replace(f"{chart_sheet_name}!", "").replace('$', "")
                chart_cat_range = chart_cat_range.split(",")[0] + ":" + chart_cat_range.split(",")[-1] if "," in chart_cat_range else chart_cat_range
                cat_data = []
                for element in wb[chart_sheet_name][chart_cat_range]:
                    for sub_element in element:
                        if type(sub_element) == Cell:
                            cat_data.append(sub_element.value)

                chart_sheet_name = input_val_range.replace('(', '').replace(')', '').replace("'", "").split(sep="!")[0]

                chart_val_range = input_val_range.replace('(', '').replace(')', '').replace("'", "").replace(f"{chart_sheet_name}!", "")
                chart_val_range = chart_val_range.split(",")[0] + ":" + chart_val_range.split(",")[-1] if "," in chart_val_range else chart_val_range
                val_data = []
                for element in wb[chart_sheet_name][chart_val_range]:
                    for sub_element in element:
                        if type(sub_element) == Cell:
                            val_data.append(sub_element.value)
                series_name = serie.tx.v if serie.tx else None

                ser = {
                    "category_axis_data": cat_data,
                    "value_axis_data": val_data,
                    "category_value_format": input_cat_format,
                    "values_value_format": input_val_format,
                    "series_name": series_name if series_name else None
                }
                series.append(ser)

            # Generate the figures
            chart_data = {
                "chart_title": chart_title,
                "chart_type": chart_type,
                "x_axis_title": x_axis_title,
                "y_axis_title": y_axis_title,
                "series": series,
            }

            figure_data = self.create_ploty_figure(chart_data)
            figures.append(figure_data)

        # Close the workbook
        wb.close()

        return figures

    def validate_sheet_names(self):
        """Validate that the input sheet and output sheets are present"""
        wb = self.workbook
        if not all(sheetname in wb for sheetname in ["viktor-input-sheet", "viktor-output-sheet"]):
            os.unlink(self.excel_file_path)
            raise UserError(
                "The sheet names are not correctly formatted.",
                input_violations=[
                    InputViolation(message="Please check the sheet and follow the documentation", fields=["excel_file"])
                ],
            )

    def get_figure_titles(self):
        """Generate dict with all the names of each figure to include in app template"""
        figure_list = []

        for i, chart in enumerate(self.charts):
            # Get the chart titles
            if chart.title:
                chart_title = ''.join([title_element.t for title_element in chart.title.tx.rich.p[0].r])
            else:
                chart_title = f"Untitled Chart {i}"
            clean_name = [s.lower() for s in chart_title.replace(" ", "_") if s.isalnum() or s == "_"]
            figure_name = "".join(clean_name)
            figure_type = chart.tagname
            figure_list.append(
                {
                    "name": chart_title,
                    "concat_name": figure_name,
                    "type": figure_type,
                }
            )

        return figure_list

    @staticmethod
    def create_ploty_figure(chart_data: dict):
        """Creates ploty figure based on the extracted chart data"""
        fig = go.Figure()
        if chart_data["chart_type"] == "lineChart":
            for ser in chart_data["series"]:
                fig.add_trace(go.Scatter(x=ser["category_axis_data"], y=ser["value_axis_data"], mode='lines', name=ser["series_name"]))
            fig.update_layout(
                title_text=chart_data["chart_title"],
                xaxis_title=chart_data["x_axis_title"],
                yaxis_title=chart_data["y_axis_title"],
                yaxis_tickformat=chart_data["series"][0]["values_value_format"],
                xaxis_tickformat=chart_data["series"][0]["category_value_format"],
            )
        if chart_data["chart_type"] == "barChart":
            for ser in chart_data["series"]:
                fig.add_trace(go.Bar(x=ser["category_axis_data"], y=ser["value_axis_data"]))
            fig.update_layout(
                title_text=chart_data["chart_title"],
                xaxis_title=chart_data["x_axis_title"],
                yaxis_title=chart_data["y_axis_title"],
                yaxis_tickformat=chart_data["series"][0]["values_value_format"],
                xaxis_tickformat=chart_data["series"][0]["category_value_format"],
            )
        if chart_data["chart_type"] == "pieChart":
            for ser in chart_data["series"]:
                fig.add_trace(go.Pie(labels=ser["category_axis_data"], values=ser["value_axis_data"]))
            fig.update_layout(
                title_text=chart_data["chart_title"],
            )
        if chart_data["chart_type"] == "scatterChart":
            for ser in chart_data["series"]:
                fig.add_trace(go.Scatter(x=ser["category_axis_data"], y=ser["value_axis_data"]))
            fig.update_layout(
                title_text=chart_data["chart_title"],
                xaxis_title=chart_data["x_axis_title"],
                yaxis_title=chart_data["y_axis_title"],
                yaxis_tickformat=chart_data["series"][0]["values_value_format"],
                xaxis_tickformat=chart_data["series"][0]["category_value_format"],
            )

        chart_data["fig"] = fig
        return chart_data
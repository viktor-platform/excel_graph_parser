from io import BytesIO

from openpyxl import load_workbook, Workbook
from openpyxl.cell import Cell
from viktor import File
from viktor.external.spreadsheet import SpreadsheetCalculation

import plotly.graph_objects as go

ALLOWED_CHART_TYPES = ["lineChart", "scatterChart", "barChart", "pieChart"]


class ExcelChartParser:
    """ Extract charts from Excel sheets that are converted to a Plotly format.

    Currently, the following chart types are supported:

    - barChart
    - lineChart
    - pieChart
    - scatterChart

    Example usage:

    ... code-block:: python

        spreadsheet = SpreadsheetCalculation(...)
        parser = ExcelChartParser(spreadsheet)
        fig = parser.get_plotly_figure("My Chart")

    """

    def __init__(self, spreadsheet_calculation: SpreadsheetCalculation):
        """
        :param spreadsheet_calculation: input spreadsheet.
        """
        file = spreadsheet_calculation._file
        if isinstance(file, File):
            with file.open_binary() as r:
                self.workbook = load_workbook(filename=r, data_only=True)
        elif isinstance(file, BytesIO):
            self.workbook = load_workbook(filename=file, data_only=True)
        else:
            raise NotImplementedError

        self._spreadsheet_calculation = spreadsheet_calculation

        # Gather charts by looping through sheets
        self._charts_map = {}
        untitled_index = 1
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            for chart in sheet._charts:  # could be empty
                if chart.title:
                    chart_title = ''.join([title_element.t for title_element in chart.title.tx.rich.p[0].r])
                else:
                    chart_title = f"Untitled {untitled_index}"
                    untitled_index += 1

                self._charts_map[chart_title] = chart

    def get_plotly_figure(self, chart_title: str) -> go.Figure:
        """Gets chart by title and returns it as Plotly figure."""
        if chart_title not in self._charts_map:
            raise ValueError(f"No chart found with title: {chart_title}")

        spreadsheet = self._spreadsheet_calculation
        result = spreadsheet.evaluate(include_filled_file=True)
        wb = load_workbook(BytesIO(result.file_content), data_only=True)
        try:
            chart_data = self._parse_chart_data(chart_title, wb)
        finally:
            wb.close()

        return self._create_plotly_figure(chart_data)

    def _parse_chart_data(self, chart_title: str, wb: Workbook) -> dict:
        """Extracts chart data from the provided workbook"""
        chart = self._charts_map[chart_title]

        # Get the general chart elements
        chart_type = chart.tagname
        if chart_type not in ALLOWED_CHART_TYPES:
            raise TypeError(
                f"Chart '{chart_title}' (type {chart_type}) cannot be parsed. Allowed types are: {ALLOWED_CHART_TYPES}"
            )

        x_axis_title, y_axis_title = None, None
        if chart_type != "pieChart":
            if chart.x_axis.title:
                x_axis_title = chart.x_axis.title.tx.rich.p[0].r[-1].t
            if chart.y_axis.title:
                y_axis_title = chart.y_axis.title.tx.rich.p[0].r[-1].t

        # Get series data
        series = []
        for serie in chart.series:
            if chart_type == "scatterChart":
                if serie.xVal:
                    if serie.xVal.strRef:
                        input_cat_range = serie.xVal.strRef.f
                        input_cat_format = None
                    elif serie.xVal.numRef:
                        input_cat_range = serie.xVal.numRef.f
                        input_cat_format = serie.xVal.numRef.numCache.formatCode
                        input_cat_format = None if input_cat_format == "General" else input_cat_format

                input_val_range = serie.yVal.numRef.f
                input_val_format = serie.yVal.numRef.numCache.formatCode
                input_val_format = None if input_val_format == "General" else input_val_format

            else:
                if serie.cat:
                    # if no category data in the sequence, use the one that was set for the previous sequence
                    if serie.cat.strRef:
                        input_cat_range = serie.cat.strRef.f
                        input_cat_format = None
                    elif serie.cat.numRef:
                        input_cat_range = serie.cat.numRef.f
                        input_cat_format = serie.cat.numRef.numCache.formatCode
                        input_cat_format = None if input_cat_format == "General" else input_cat_format

                input_val_range = serie.val.numRef.f
                input_val_format = serie.val.numRef.numCache.formatCode
                input_val_format = None if input_val_format == "General" else input_val_format

            chart_sheet_name = input_cat_range.replace('(', '').replace(')', '').replace("'", "").split(sep="!")[0]

            chart_cat_range = input_cat_range.replace('(', '').replace(')', '').replace("'", "").replace(f"{chart_sheet_name}!", "").replace('$', "")
            chart_cat_range = chart_cat_range.split(",")[0] + ":" + chart_cat_range.split(",")[-1] if "," in chart_cat_range else chart_cat_range
            cat_data = []
            for element in wb[chart_sheet_name][chart_cat_range]:
                for sub_element in element:
                    if type(sub_element) == Cell:
                        cat_data.append(sub_element.value)

            chart_sheet_name = input_val_range.replace('(', '').replace(')', '').replace("'", "").split(sep="!")[0]

            chart_val_range = input_val_range.replace('(', '').replace(')', '').replace("'", "").replace(f"{chart_sheet_name}!", "")
            chart_val_range = chart_val_range.split(",")[0] + ":" + chart_val_range.split(",")[-1] if "," in chart_val_range else chart_val_range
            val_data = []
            for element in wb[chart_sheet_name][chart_val_range]:
                for sub_element in element:
                    if type(sub_element) == Cell:
                        val_data.append(sub_element.value)

            series_name = serie.tx.v if serie.tx else None
            ser = {
                "category_axis_data": cat_data,
                "value_axis_data": val_data,
                "category_value_format": input_cat_format,
                "values_value_format": input_val_format,
                "series_name": series_name if series_name else None
            }
            series.append(ser)

        chart_data = {
            "chart_title": chart_title,
            "chart_type": chart_type,
            "x_axis_title": x_axis_title,
            "y_axis_title": y_axis_title,
            "series": series,
        }

        return chart_data

    @staticmethod
    def _create_plotly_figure(chart_data: dict) -> go.Figure:
        """Creates plotly figure based on the extracted chart data"""
        fig = go.Figure()
        if chart_data["chart_type"] == "lineChart":
            for ser in chart_data["series"]:
                fig.add_trace(go.Scatter(
                    x=ser["category_axis_data"],
                    y=ser["value_axis_data"],
                    mode='lines',
                    name=ser["series_name"]
                ))
            fig.update_layout(
                title_text=chart_data["chart_title"],
                xaxis_title=chart_data["x_axis_title"],
                yaxis_title=chart_data["y_axis_title"],
                yaxis_tickformat=chart_data["series"][0]["values_value_format"],
                xaxis_tickformat=chart_data["series"][0]["category_value_format"],
            )
        if chart_data["chart_type"] == "barChart":
            for ser in chart_data["series"]:
                fig.add_trace(go.Bar(x=ser["category_axis_data"], y=ser["value_axis_data"]))
            fig.update_layout(
                title_text=chart_data["chart_title"],
                xaxis_title=chart_data["x_axis_title"],
                yaxis_title=chart_data["y_axis_title"],
                yaxis_tickformat=chart_data["series"][0]["values_value_format"],
                xaxis_tickformat=chart_data["series"][0]["category_value_format"],
            )
        if chart_data["chart_type"] == "pieChart":
            for ser in chart_data["series"]:
                fig.add_trace(go.Pie(labels=ser["category_axis_data"], values=ser["value_axis_data"]))
            fig.update_layout(
                title_text=chart_data["chart_title"],
            )
        if chart_data["chart_type"] == "scatterChart":
            for ser in chart_data["series"]:
                fig.add_trace(go.Scatter(x=ser["category_axis_data"], y=ser["value_axis_data"]))
            fig.update_layout(
                title_text=chart_data["chart_title"],
                xaxis_title=chart_data["x_axis_title"],
                yaxis_title=chart_data["y_axis_title"],
                yaxis_tickformat=chart_data["series"][0]["values_value_format"],
                xaxis_tickformat=chart_data["series"][0]["category_value_format"],
            )

        return fig
